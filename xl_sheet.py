import openpyxl as xl
from openpyxl.chart import BarChart,Reference 
# we have a excel sheet which have a mistaken price field. corrected values should be lesser by 10 %. so here we are trying to make in correct by multiplying each price field by 0.9 which will decrease 10% from it.
def process_workbook(filename):
      work_book = xl.load_workbook(filename)
      sheet = work_book["Sheet1"]
      row_count = sheet.max_row     

# iterate through all rows and multiply the value by 0.9
      for row in range(2,row_count+1):
      # we are ignoring 1st row ans starting from 2 upto 4 
            cell =  sheet.cell(row,3) #row,coumn
            correacted_price = cell.value * 0.9
            corrected_price_cell = cell
            corrected_price_cell.value = correacted_price

      values= Reference(sheet,min_row=2,max_row=row_count,min_col=4,max_col=4) 
      chart = BarChart()
      chart.add_data(values)
      sheet.add_chart(chart,"e2")
      work_book.save(filename)
      